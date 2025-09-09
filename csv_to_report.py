#!/usr/bin/env python3
"""
CSV to Excel Report Generator
============================

This script processes measure data from CSV files and JSON profile to generate
an Excel report with historical values, latest values, scores, and summary.

Requirements:
- measure_value.csv: Historical values for each measure
- measure_score.csv: Scores for each measure  
- measure_profile.json: Measure metadata (name, unit)

Output:
- report_output.xlsx: Excel file with Report and Summary sheets
"""

import pandas as pd
import json
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Tuple
import os
from datetime import datetime
import html as ihtml
from typing import Dict, List, Optional

class CSVToReportGenerator:
    """Generate Excel report from CSV and JSON files"""
    
    def __init__(self, value_file: str, score_file: str, measure_profile_file: str, categories: dict,
                  frequency: str = 'M'):
        """
        Initialize the generator with input files
        
        Args:
            value_file: Path to measure_value.csv
            score_file: Path to measure_score.csv  
            measure_profile_file: Path to measure_profile.json
            categories: Dictionary of measure categories
            display_period: Tuple of (start_date, end_date) as 'YYYY/MM/DD' strings, default: 去年底~今日
            frequency: Data frequency, e.g. '月', default: '月'
        """
        self.value_file = value_file
        self.score_file = score_file
        self.measure_profile_file = measure_profile_file
        self.categories = categories

        self.frequency = frequency
        
    def clean_column_name(self, col_name: str) -> str:
        """Clean column names by removing extra spaces and newlines"""
        return col_name.strip().replace('\n', '')

    def load_data(self) -> Tuple[pd.DataFrame, pd.DataFrame, Dict]:
        """
        Load and clean data from input files
        
        Returns:
            Tuple of (value_df, score_df, measure_profile_dict)
        """
        # Load CSV files with big5 encoding
        value_df = pd.read_csv(self.value_file, encoding='big5')
        score_df = pd.read_csv(self.score_file, encoding='big5')
        # Ensure 'Date' column is present
        if 'Date' not in value_df.columns or 'Date' not in score_df.columns:
            raise ValueError("Both CSV files must contain a 'Date' column.")
        #Date column should be in datetime format
        value_df['Date'] = pd.to_datetime(value_df['Date'])
        score_df['Date'] = pd.to_datetime(score_df['Date'])

        # Clean column names
        value_df.columns = [self.clean_column_name(col) for col in value_df.columns]
        score_df.columns = [self.clean_column_name(col) for col in score_df.columns]
        
        # Load JSON measure_profile with utf-8 encoding
        with open(self.measure_profile_file, 'r', encoding='utf-8') as f:
            measure_profile = json.load(f)
            
        return value_df, score_df, measure_profile
    
    def get_measure_category(self, measure_id: str) -> str:
        """Get the category of a measure"""
        for category, measures in self.categories.items():
            if measure_id in measures:
                return category
        return "未分類"
    
    def create_report_sheet(self, 
                            display_period: tuple,
                            value_df: pd.DataFrame, 
                            score_df: pd.DataFrame, 
                            measure_profile: Dict) -> pd.DataFrame:
        """
        Create the main report dataframe with proper column layout
        
        Args:
            display_period: Tuple of (start_date, end_date) as 'YYYY/MM/DD' strings
            value_df: DataFrame with historical values
            score_df: DataFrame with scores
            measure_profile: Measure profile dictionary
            
        Returns:
            DataFrame ready for display
        """
        report_data = []
        # filter value_df and score_df by display_period
        start_date, end_date = pd.to_datetime(display_period[0]), pd.to_datetime(display_period[1])
        value_df = value_df[(pd.to_datetime(value_df['Date']) >= start_date) & (pd.to_datetime(value_df['Date']) <= end_date)]
        score_df = score_df[(pd.to_datetime(score_df['Date']) >= start_date) & (pd.to_datetime(score_df['Date']) <= end_date)]

        # Get date columns for historical values
        date_columns = value_df['Date'].tolist()
        
        # Process each measure
        measure_columns = [col for col in value_df.columns if col != 'Date']
        
        for measure_id in measure_columns:
            if measure_id not in measure_profile:
                continue               
           
            # Get historical values
            historical_values = value_df[measure_id].tolist()            
           
            row = {
                'category': self.get_measure_category(measure_id),
                'measure_name': measure_profile[measure_id]['name'],
                'unit': measure_profile[measure_id].get('unit', ''),
            }

            for date, value in zip(date_columns, historical_values):
                row[datetime.strftime(date, '%Y-%m-%d')] = value

            row['score'] = score_df[measure_id].iloc[-1]                
            report_data.append(row)
        

        # Convert to DataFrame
        report_df = pd.DataFrame(report_data)
        # sum of scores by category
        report_df['score_total'] = report_df.groupby('category')['score'].transform('sum')

        # Sort by category (as in categories.keys()) and then by the order in each category's list
        category_order = {cat: i for i, cat in enumerate(self.categories.keys())}
        measure_order = {}
        for cat, measures in self.categories.items():
            for i, measure in enumerate(measures):
                measure_order[measure] = i  
        report_df['category_order'] = report_df['category'].map(category_order).fillna(999)
        report_df['measure_order'] = report_df['measure_name'].map(measure_order).fillna(999)
        report_df = report_df.sort_values(by=['category_order', 'measure_order'])
        report_df = report_df.drop(columns=['category_order', 'measure_order'])

        return report_df


    def export_to_html(self, df: pd.DataFrame, filename: str = "report.html", 
                       add_style: bool = True, 
                       rename: Optional[Dict[str, str]] = None,
                       col_mergecell: Optional[List[str]] = None) -> str:
        """
        將 DataFrame 輸出為 HTML 表格，並針對指定欄位進行儲存格列合併(rowspan)。

        參數：
            df            : 來源 DataFrame（會依目前順序判斷相鄰合併）
            filename      : 輸出 HTML 檔名
            add_style     : 是否插入內嵌 CSS 樣式
            rename        : 欄位重命名字典, e.g. {'category': '類別', 'measure_name': '指標名稱'}
            col_mergecell : 要進行合併的欄位清單, e.g. ['category', 'score_total']

        回傳：
            產出的 HTML 字串（亦會寫入 filename）
        """
        # 製作副本並應用重命名
        _df = df.copy()
        if rename:
            _df = _df.rename(columns=rename)
        
        # 設定參數
        table_id = "report-table"
        merge_cols = col_mergecell if col_mergecell else []

        
        # 為了相等性判斷一致，先把 NaN 轉為空字串
        for col in _df.columns:
            _df[col] = _df[col].fillna("")

        # --- 計算連續區段的 rowspan 起點與長度 ---
        def compute_runs(series: pd.Series) -> Dict[int, int]:
            """回傳 {起始列索引: 連續長度} 只記錄每段起始列"""
            runs = {}
            i, n = 0, len(series)
            while i < n:
                j = i + 1
                while j < n and series.iloc[j] == series.iloc[i]:
                    j += 1
                runs[i] = j - i
                i = j
            return runs

        # 計算每個要合併的欄位的連續區段
        merge_runs = {}
        for col_name in merge_cols:
            # 檢查原始欄名是否存在於原始 DataFrame
            if col_name in df.columns:
                # 取得重命名後的欄名
                renamed_col = rename.get(col_name, col_name) if rename else col_name
                if renamed_col in _df.columns:
                    merge_runs[renamed_col] = compute_runs(_df[renamed_col])

        # --- 準備輸出欄順序：維持重命名後的 df 欄位順序 ---
        columns: List[str] = list(_df.columns)

        # --- 內嵌樣式（可關閉或自行覆寫） ---
        style_block = ""
        if add_style:
            style_block = f"""
    <style>
    table#{table_id} {{
        border-collapse: collapse;
        width: 100%;
        table-layout: fixed;
        word-wrap: break-word;
    }}
    table#{table_id} th, table#{table_id} td {{
        border: 1px solid #ddd;
        padding: 6px 8px;
        vertical-align: middle;
        text-align: center;
    }}
    table#{table_id} thead th {{
        background: #f4f6f8;
        font-weight: 600;
    }}
    table#{table_id} tbody tr:nth-child(even) {{
        background: #fafafa;
    }}
    /* sticky header（如不需要可刪除） */
    table#{table_id} thead th {{
        position: sticky;
        top: 0;
        z-index: 2;
    }}
    </style>
    """

        # --- 建立表頭 ---
        thead_cells = "".join(f"<th>{ihtml.escape(str(col))}</th>" for col in columns)
        thead_html = f"<thead><tr>{thead_cells}</tr></thead>"

        # --- 建立表身（處理 rowspan） ---
        body_rows_html = []
        nrows = len(_df)

        for r in range(nrows):
            tds = []

            for col in columns:
                col_index = _df.columns.get_loc(col)
                val = _df.iloc[r, col_index]
                txt = "" if pd.isna(val) else str(val)
                esc = ihtml.escape(txt)

                # 檢查是否為需要合併的欄位
                if col in merge_runs:
                    # 只有段落起點才輸出 <td rowspan="...">
                    if r in merge_runs[col]:
                        rs = merge_runs[col][r]
                        tds.append(f'<td rowspan="{rs}">{esc}</td>')
                    else:
                        # 非起點列就跳過（由起點列的 rowspan 覆蓋）
                        pass
                else:
                    tds.append(f"<td>{esc}</td>")

            body_rows_html.append("<tr>" + "".join(tds) + "</tr>")

        tbody_html = "<tbody>\n" + "\n".join(body_rows_html) + "\n</tbody>"

        html = f"""<!DOCTYPE html>
    <html lang="zh-Hant">
    <head>
    <meta charset="utf-8">
    <title>報表</title>
    {style_block}
    </head>
    <body>
    <table id="{table_id}">
    {thead_html}
    {tbody_html}
    </table>
    </body>
    </html>
    """

        with open(filename, "w", encoding="utf-8") as f:
            f.write(html)

        return html


    def export_to_excel(self, df: pd.DataFrame, filename: str = "output.xlsx", 
                        rename: Optional[Dict[str, str]] = None, 
                        col_mergecell: Optional[List[str]] = None):
        """
        Export DataFrame to Excel with optional header renaming and configurable cell merging
        
        Args:
            df: DataFrame to export
            filename: Output Excel filename
            rename: Dictionary to rename column headers, e.g. {'category': '類別', 'measure_name': '指標名稱'}
            col_mergecell: List of column names to apply cell merging, e.g. ['category', 'score_total']
        """
        # 製作副本以避免修改原 DataFrame
        output_df = df.copy()
        
        # 應用列名重命名
        if rename:
            output_df = output_df.rename(columns=rename)
        
        # 先輸出到 Excel
        output_df.to_excel(filename, index=False)

        # 如果沒有指定合併欄位，則不進行合併
        if not col_mergecell:
            print(f"已輸出完成：{filename}")
            return

        # 讀取 Excel，進行合併
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # 找出要合併的欄位位置
        merge_columns = []
        for col_name in col_mergecell:
            # 檢查原始欄名是否存在
            if col_name in df.columns:
                # 取得重命名後的欄名
                renamed_col = rename.get(col_name, col_name) if rename else col_name
                if renamed_col in output_df.columns:
                    col_index = output_df.columns.get_loc(renamed_col) + 1  # 轉為 Excel index (從 1 開始)
                    merge_columns.append(col_index)

        start_row = 2  # 第一列是標題，從第二列開始
        for col in merge_columns:
            prev_val = ws.cell(row=start_row, column=col).value
            merge_start = start_row

            for row in range(start_row + 1, ws.max_row + 1):
                current_val = ws.cell(row=row, column=col).value
                if current_val != prev_val:
                    # 合併上一段
                    if merge_start < row - 1:
                        ws.merge_cells(
                            start_row=merge_start,
                            start_column=col,
                            end_row=row - 1,
                            end_column=col
                        )
                    merge_start = row
                    prev_val = current_val

            # 處理最後一段
            if merge_start < ws.max_row:
                ws.merge_cells(
                    start_row=merge_start,
                    start_column=col,
                    end_row=ws.max_row,
                    end_column=col
                )

        wb.save(filename)
        print(f"已輸出並合併完成：{filename}")


    def adjust_df_frequency(self, df: pd.DataFrame, date_col: str = 'Date', frequency: str = 'M') -> pd.DataFrame:
        """Adjust DataFrame to specified frequency by resampling"""
        df[date_col] = pd.to_datetime(df[date_col])
        df = df.set_index(date_col).resample(frequency).last().reset_index()
        return df
        
    def generate_report(self, display_period: tuple, output_file: str = "report_output.html"):
        """
        Generate the complete Html report
        
        Args:
            display_period: Tuple of (start_date, end_date) as 'YYYY/MM/DD' strings
            output_file: Output html file path
        """
        print("Loading data...")
        value_df, score_df, measure_profile = self.load_data()

        #調整資料
        value_df = self.adjust_df_frequency(value_df, frequency=self.frequency)
        score_df = self.adjust_df_frequency(score_df, frequency=self.frequency)

        print("Creating report sheet...")
        # Create the report DataFrame
        report_df = self.create_report_sheet(display_period, value_df, score_df, measure_profile)

        print("exporting to HTML...")
        # 使用預設的合併欄位以維持向後相容性
        default_col_mergecell = ['category', 'score_total']
        self.export_to_html(report_df, output_file, col_mergecell=default_col_mergecell)

        print("Exporting to Excel...")
        # 使用預設的合併欄位以維持向後相容性
        self.export_to_excel(report_df, output_file.replace(".html", ".xlsx"), col_mergecell=default_col_mergecell)

def main():
    """Main function to run the report generator"""
    
    # Input
    display_period = ('2024/12/1','2025/7/31')
    value_file = "measure_value.csv"
    score_file = "measure_score.csv"
    measure_profile_file = "measure_profile.json"
    categories = {
            "總經面指標": [
                "台灣領先指標_id", "ISM製造業指數_id", "台灣外銷訂單_id", 
                "台灣工業生產指數_id", "台灣貿易收支_id", "台灣零售銷售_id", 
                "台灣失業率_id", "台灣CPI_id", "台灣M1B-M2_id"
            ],
            "技術面指標": [
                "加權指數乖離率_id", "OTC指數乖離率_id", "加權指數MACD_id", 
                "OTC指數MACD_id", "加權指數DIF_id", "加權指數ADX_id"
            ],
            "評價面指標": [
                "加權指數本益比_id", "台灣50指數本益比_id", "台灣中型100指數本益比_id", 
                "台灣高股息指數本益比_id", "OTC指數本益比_id", "加權指數股價淨值比_id", 
                "台灣50指數股價淨值比_id", "台灣中型100指數股價淨值比_id", 
                "台灣高股息指數股價淨值比_id", "OTC指數股價淨值比_id"
            ]
        }
    # Check if files exist
    for file_path in [value_file, score_file, measure_profile_file]:
        if not os.path.exists(file_path):
            print(f"Error: File not found: {file_path}")
            return
    
    # Generate report
    generator = CSVToReportGenerator(value_file, score_file, measure_profile_file, categories)
    generator.generate_report(display_period)


if __name__ == "__main__":
    main()