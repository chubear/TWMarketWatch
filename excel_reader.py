#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel data reader module for TWMarketWatch
Reads specific ranges from Excel worksheets and converts them to web-displayable format
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Tuple, Optional


class ExcelDataReader:
    """Class to handle reading Excel data for web display"""
    
    def __init__(self, excel_file_path: str):
        """Initialize with path to Excel file"""
        self.excel_file_path = excel_file_path
        
    def read_worksheet_range(self, worksheet_name: str, range_str: str) -> pd.DataFrame:
        """
        Read a specific range from a worksheet
        
        Args:
            worksheet_name: Name of the worksheet
            range_str: Excel range like "A1:X61"
            
        Returns:
            DataFrame with the data from the specified range
        """
        try:
            # Use openpyxl to read the specific range
            wb = openpyxl.load_workbook(self.excel_file_path, data_only=True)
            
            if worksheet_name not in wb.sheetnames:
                raise ValueError(f"Worksheet '{worksheet_name}' not found in file")
            
            ws = wb[worksheet_name]
            
            # Parse the range (e.g., "A1:X61")
            start_cell, end_cell = range_str.split(':')
            
            # Get data from the range
            data = []
            for row in ws[range_str]:
                row_data = []
                for cell in row:
                    value = cell.value
                    if value is None:
                        value = ""
                    row_data.append(value)
                data.append(row_data)
            
            wb.close()
            
            # Convert to DataFrame
            if data:
                df = pd.DataFrame(data)
                # Clean up column names if first row contains headers
                if len(data) > 0:
                    # Use first row as potential headers, but keep as numeric indices for now
                    pass
                return df
            else:
                return pd.DataFrame()
                
        except Exception as e:
            print(f"Error reading worksheet range: {e}")
            return pd.DataFrame()
    
    def get_investment_decision_data(self) -> pd.DataFrame:
        """Get data from 國內股市當年度投資決策建議表 worksheet (A1:X61)"""
        worksheet_name = "國內股市當年度投資決策建議表 "  # Note the trailing space
        return self.read_worksheet_range(worksheet_name, "A1:X61")
    
    def get_investment_analysis_data(self) -> pd.DataFrame:
        """Get data from 國外股市當年度投資評等分析表 worksheet (A1:X55)"""
        worksheet_name = "國外股市當年度投資評等分析表 "  # Note the trailing space
        return self.read_worksheet_range(worksheet_name, "A1:X55")
    
    def get_all_data(self) -> Dict[str, pd.DataFrame]:
        """Get all required data as a dictionary"""
        return {
            'investment_decision': self.get_investment_decision_data(),
            'investment_analysis': self.get_investment_analysis_data()
        }
    
    def df_to_html_table(self, df: pd.DataFrame, table_id: str = "", 
                        table_class: str = "table table-striped") -> str:
        """
        Convert DataFrame to HTML table
        
        Args:
            df: DataFrame to convert
            table_id: HTML id for the table
            table_class: CSS classes for the table
            
        Returns:
            HTML string representation of the table
        """
        if df.empty:
            return "<p>No data available</p>"
        
        # Convert DataFrame to HTML with custom styling
        html = df.to_html(
            classes=table_class,
            table_id=table_id,
            escape=False,
            index=False,
            border=0
        )
        
        # Replace NaN values with empty strings in the HTML
        html = html.replace('NaN', '')
        html = html.replace('nan', '')
        html = html.replace('None', '')
        
        return html


def main():
    """Test the ExcelDataReader"""
    excel_file = "01.股市經濟及市場指標2025Q2.xlsx"
    reader = ExcelDataReader(excel_file)
    
    print("Testing ExcelDataReader...")
    
    # Test reading investment decision data
    decision_data = reader.get_investment_decision_data()
    print(f"\nInvestment Decision Data shape: {decision_data.shape}")
    if not decision_data.empty:
        print("First few rows:")
        print(decision_data.head())
    
    # Test reading investment analysis data
    analysis_data = reader.get_investment_analysis_data()
    print(f"\nInvestment Analysis Data shape: {analysis_data.shape}")
    if not analysis_data.empty:
        print("First few rows:")
        print(analysis_data.head())


if __name__ == "__main__":
    main()