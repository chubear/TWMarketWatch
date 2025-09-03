#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to explore the Excel file structure
"""

import pandas as pd
import openpyxl

def explore_excel_file():
    excel_file = "01.股市經濟及市場指標2025Q2.xlsx"
    
    print(f"Exploring Excel file: {excel_file}")
    
    # Load workbook to get worksheet names
    wb = openpyxl.load_workbook(excel_file)
    print(f"\nWorksheets in the file:")
    for i, sheet_name in enumerate(wb.sheetnames):
        print(f"{i+1}. {sheet_name}")
    
    # Check if our target worksheets exist
    # Note: Based on the output, the actual sheet names have trailing spaces
    # Also checking if it's "國內" or "國外" for the second sheet
    target_sheets = ["國內股市當年度投資決策建議表 ", "國外股市當年度投資評等分析表 ", "國內股市當年度投資評等分析表"]
    
    for sheet_name in target_sheets:
        if sheet_name in wb.sheetnames:
            print(f"\n✓ Found target worksheet: '{sheet_name}'")
            
            # Get the worksheet
            ws = wb[sheet_name]
            
            # Get dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"  Dimensions: {max_row} rows x {max_col} columns")
            
            # Preview first few rows and columns
            print(f"  Preview of data:")
            for row in range(1, min(6, max_row + 1)):
                row_data = []
                for col in range(1, min(6, max_col + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is None:
                        cell_value = ""
                    row_data.append(str(cell_value)[:20])  # Truncate long values
                print(f"    Row {row}: {row_data}")
        else:
            print(f"\n✗ Target worksheet NOT found: '{sheet_name}'")
    
    wb.close()

if __name__ == "__main__":
    explore_excel_file()